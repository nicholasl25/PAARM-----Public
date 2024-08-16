from .Predictor import Predictor
from .Helper import check_count, standard_dev, get_title, get_lf, current_month, start_year
from statsmodels.base.model import LikelihoodModel
import pandas as pd
import numpy as np
from datetime import datetime
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

"""# Defines Class Past which inherits from Predictor"""

# Class that functions are used when "Backtest" and "Short Term" are selected in Settings.html
class Past(Predictor):
    # class defintion - see Predictor's __init__() method
    def __init__(self, models, weights, matplot, params, path, filter, covid):
        super().__init__(models, weights, matplot, params, path, filter, covid)

    # Returns None -- calls other functions in Past which export graphs/numbers/params
    # time: int representing how many months the prediction should start after current_month
    # data: all datasets contained as a dictionary where the keys are in ["type", "airport", "airline"] format and the
    # values are csvs with a "Total Pax" and "time" column -- DATA FOR PAX
    # exog: all exogenous variables contained as a dictionary where the keys are in ["type", "airport", "airline"] format and the
    # values are csvs with a "Total Seats" and "time" column -- DATA FOR SEATS
    # boole2: boolean that is True if optimistic/pessimistic values should be included in the report and False otherwise
    # name: string folder name of the report in the /Reports folder
    # global_exo: all exogenous variables that should be considered for all datasets may be None. Must be a pandas dataframe 
    # that has a time column aligned with the time column of values in data (usually GDP)
    def predict(self, time, data, exog, exog2, boole2, name, global_exo):
        dct = {}
        ind = -1 * Past.max_iter # Max number of models to be trained (default 1000 from Predictor)
        for key in data.keys():
            include = self.filter.filter_keys(key)
            if ind < 0 and data[key] is not None and key in exog.keys() and check_count(data[key]) and \
            check_count(exog[key]) and len(data[key]["time"]) > time and include:
                assert len(data[key]["time"]) <= 2*time, "NICK SAYS: A time value given was too low, times should be at least half of data length"
                assert len(exog[key]["time"]) >= time, "NICK SAYS: A time value given is out of range of exogenous variables"
                end = len(data[key]["time"])
                dct = {key: [] for key in range(time, int(end))}
                merged = pd.merge(data[key], exog[key], on='time', how='inner')
                merged2 = pd.merge(merged, global_exo, on = "time", how = "left")[Past.exog_targets] \
                if global_exo is not None else merged[["Total Seats", "time"]]
                # include Covid as an exogenous variable that is 1 between March 2020 and Dec 2020 and 0 elsewhere
                if self.covid.get_include():
                    merged2["Covid"] = merged2["time"].apply(self.covid.get_function())
                merged2 = merged2.drop(columns = ["time"])
                count = 1
                for model in self.models:
                    ind += 1
                    mod = model(merged[:time][Past.target], merged2[:time])
                    model_fit = mod.fit()
                    # Export model parameters if recquired
                    if self.params:
                        self.get_model_params(model_fit, self.path, name, time, list(key), count)

                    # Make prediction (accounting for several model types)
                    if isinstance(mod, LikelihoodModel):
                        pred = model_fit.predict(time, end-1, exog = merged2[time:])
                    else:
                        forecast = model_fit.forecast(start=time, horizon=end-1, x=None)
                        pred = forecast.mean.stack().reset_index(drop=True)

                    for index, value in enumerate(pred):
                        dct[index+time].append(value)
                    count += 1

                # Make report directory
                if self.path is not None:
                    if not os.path.exists(self.path + name + "/"):
                        os.makedirs(self.path + name + "/")

                # Generates matplotlib graphs of all the data
                if self.matplot:
                    self.save_graphs(data[key], dct, time, list(key), self.path, self.weights, name, boole2)

                # Generates the excel file containing all the date (boole2 is True if range of values should be reported)
                if boole2:
                    self.save_numbers(dct, data[key], time, list(key), self.path, exog[key]["Total Seats"], name, self.weights)
                else:
                    self.save_numbers2(dct, data[key], time, list(key), self.path, exog[key]["Total Seats"], name, self.weights)


    # Exports a text file of the model parameters for every model created in your report directory
    # model: trained model of which the parameters are exported
    # path: directory path the report should be saved into (string)
    # time: number representing how many months are being predicted
    # name: name of report to be saved in directory (string)
    # keys: list of strings that represent [type, airport, airline]
    # number: integer that represents what number the model is in the ensemble
    # USED FOR BACKTESTING, OTHERWISE USE get_model_params in Future

    def get_model_params(self, model, path, name, time, keys, number):
    # Creates directory to save model parameters as a text file
        base_directory = path + name
        directory = os.path.join(base_directory,  self.get_date(time) + "/" + keys[0] + "/" + keys[1] + "/")
        file_name = keys[2] + " Model#" + str(number) + ".txt"
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Saves model parameters as a text file and exports them to appropriate directory
        summary_text = model.summary().as_text()
        with open(os.path.join(directory, file_name), 'w') as f:
            f.write(summary_text)


    # Generates a matplot for an ensemble models future predictions
    # x: input pandas dataframe that model was trained on
    # dct: dictionary giving giving an array of predicted values for every time key
    # time: int representing the number of months being extrapolated
    # key: list of strings that represent [type, airport, airline]
    # path: the string file path which the graphs are being saved to
    # weights: array of ints that decide how each model should be weighted
    # in the final average
    # boole2: True if optimistic/pessimistic values should be included in the graph
    # and False otherwise
    # USED FOR BACKTESTING IF NOT USE save_graphs in Future

    def save_graphs(self, x, dct, time, key, path, weights, name, boole2):
        cls = self.__class__
        mean_dict = {key: np.average(value, weights=weights, axis = 0) for key, value in dct.items()}
        mean_df = pd.DataFrame.from_dict(mean_dict, orient='index')
        plt.plot(x["time"][time:], x[cls.target][time:], label = "Actual", color = "black")
        plt.plot(x["time"][time:], mean_df, label = "Mean", color = "red")
        plt.axvline(x = time, color = 'black', ls = "--")
        if boole2:
            std_dict = {key: standard_dev(value, weights) for key, value in dct.items()}
            std_df = pd.DataFrame.from_dict(std_dict, orient='index')

            plt.fill_between(x["time"][time:], (mean_df - 2*std_df)[0], (mean_df + 2*std_df)[0],
                            color='orange', alpha=0.4, label='Confidence Interval')

            plt.fill_between(x["time"][time:], (mean_df - std_df)[0], (mean_df + std_df)[0],
                        color=(1,0.25,0), alpha=0.4, label='Confidence Interval')

        plt.suptitle(get_title(key))
        # Actual values (from time to the end)
        actual_values = x[cls.target].iloc[time:].reset_index(drop=True)
        # Forecasted values (from time to the end)
        forecasted_values = mean_df[0].reset_index(drop=True)
        # Calculate the mean absolute percentage error
        MAPE = np.mean(np.abs((actual_values - forecasted_values) / actual_values)) * 100
        plt.title("Mean Absolute Percentage Error (MAPE): " + str(round(MAPE, 2)) + "%", fontsize=10, color='red')
        plt.legend(['Actual', 'Mean Predicted'], frameon=True, loc='lower left')
        plt.ylabel('Passenger Volume')
        plt.xlabel(cls.xlabel)
        plt.grid()

        if path is not None:
            if not os.path.exists(path + name + '/' + self.get_date(time) + '/' + key[0] + "/" + key[1]):
                os.makedirs(path + name +  '/' + self.get_date(time) + '/' + key[0] + "/" + key[1])

        plt.savefig(path + name + '/' + self.get_date(time) + '/' + key[0] + "/" + key[1] + "/" + key[2] + ".png")

        mean_df.reset_index(drop=True, inplace=True)

        plt.close()


    # Helper function that generates an excel file from historic and predicted data
    # dct: dictionary giving giving an array of predicted values for every time key
    # x: pandas dataframe the ensemble was trained on (must have "time" + "Total Pax" columns)
    # time: int representing the number of months being extrapolated
    # key: list of strings that represent [type, airport, airline]
    # path: the string file path which the graphs are being saved to
    # seats: pandas dataframe of the future seat values (must have a "Total Seats" column)
    # name: string name of report that is being generated
    # weights: array of ints that decide how each model should be weighted
    # in the final average.
    # INCLUDES OPTIMISTIC/PESSIMISTIC PREDICTIONS IF NOT USE save_numbers2
    # USED FOR BACKTESTING IF YOU WANT TO EXTRAPOLATE USE save_numbers in Future

    def save_numbers(self, dct, x, time, key, path, seats, name, weights):
        cls = self.__class__
        mean_dict = {key: np.average(value, weights=weights) for key, value in dct.items()}
        std_dict = {key: standard_dev(value, weights) for key, value in dct.items()}
        opt_dict = {key: mean_dict[key] + std_dict[key] for key, value in dct.items()}
        pess_dict = {key: mean_dict[key] - std_dict[key] for key, value in dct.items()}
        real_values = x[cls.target].values[time:]
        data = {
            "time": list(range(time, len(x["time"]))),
            "Type": key[0],
            "Airport": key[1],
            "Airline": key[2],
            "Actual Pax": real_values,
            "Mean Pred Pax": list(mean_dict.values()),
            "Percent Error": np.abs(100*((np.array(real_values) - np.array(list(mean_dict.values()))) / np.array(real_values))),
            "Opt Pred Pax": list(opt_dict.values()),
            "Pes Pred Pax":  list(pess_dict.values()),
            "Total Seats": seats[time:],
            }
        df = pd.DataFrame(data)
        base_date = pd.Timestamp('2016-01-01') if cls.monthly else datetime(start_year, 1, 1)
        if cls.monthly:
            df['Date'] = df['time'].apply(lambda x: base_date + pd.DateOffset(months=x))
            df['Date'] = df['Date'].dt.strftime('%b %Y')
        else:
            df['Date'] = df['time'].apply(lambda x: base_date + pd.DateOffset(years=x))
            df['Date'] = df['Date'].dt.strftime('%Y')
        df.drop(columns = "time", inplace = True)
        df["Mean Pred Pax"] = df["Mean Pred Pax"].round().astype(int)
        df["Percent Error"] = df["Percent Error"].round(1).astype(str) + '%'
        df["Opt Pred Pax"] = df["Opt Pred Pax"].round().astype(int)
        df["Pes Pred Pax"] = df["Pes Pred Pax"].round().astype(int)
        df["Actual LF"] = get_lf(df["Actual Pax"], df["Total Seats"])
        df["Mean Pred LF"] = get_lf(df["Mean Pred Pax"], df["Total Seats"])
        df["Opt Pred LF"] = get_lf(df["Opt Pred Pax"], df["Total Seats"])
        df["Pes Pred LF"] = get_lf(df["Pes Pred Pax"], df["Total Seats"])
        # Rearrange the columns to place the date column first
        cols = ["Date", "Type", "Airport","Airline", "Actual Pax", "Mean Pred Pax",  "Opt Pred Pax", "Pes Pred Pax", "Percent Error",
                "Total Seats", "Actual LF", "Mean Pred LF", "Opt Pred LF", "Pes Pred LF"] if self.monthly else \
                 ["Date", "Type", "Airport","Airline", "Actual Pax", "Mean Pred Pax",  "Opt Pred Pax", "Pes Pred Pax", "Percent Error"]
        df = df[cols]
        df.title = get_title(key)

        sheet_name = self.get_date(time)  # One sheet per time interval

        # Load existing workbook or create a new one if it doesn't exist
        try:
            book = load_workbook(path + name + "/Data.xlsx")
            if sheet_name not in book.sheetnames:
                # Create the sheet if it doesn't exist
                book.create_sheet(sheet_name)
            with pd.ExcelWriter(path + name + "/Data.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                startrow = book[sheet_name].max_row  # Find the last row in the existing sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
        except FileNotFoundError:
            with pd.ExcelWriter(path + name + "/Data.xlsx", engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                startrow = 0  # Since this is a new sheet

        # Load workbook and select sheet to add graph
        book = load_workbook(path + name + "/Data.xlsx")
        sheet = book[sheet_name]

        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
            col_letter = col[0].column_letter  # Get the column letter
            sheet.column_dimensions[col_letter].width = 13

        # Create a line chart
        chart = LineChart()
        chart.title = get_title(key)[:-1]
        chart.style = 10
        chart.height = 6
        chart.width = 12
        chart.grouping = 'standard'
        chart.y_axis.title = "Predicted Passengers"
        chart.x_axis.title = cls.xlabel

        size = current_month - time
        # Iterate over all rows and set their height
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            sheet.row_dimensions[row[0].row].height = 150 / size if size < 10 else 15

        # Define data for the chart using only the newly added rows
        new_data_start_row = startrow + 1
        new_data_end_row = sheet.max_row
        data = Reference(sheet, min_col=5, min_row=new_data_start_row, max_col=8, max_row=new_data_end_row)
        categories = Reference(sheet, min_col=1, min_row=new_data_start_row + 1, max_row=new_data_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Set the minimum value for the y-axis
        min_value = min(df[["Pes Pred Pax", "Actual Pax"]].min())
        chart.y_axis.scaling.min = min_value

        # Add the chart to the sheet
        fst = 'P' if self.monthly else 'K'
        chart_position = fst + str(new_data_start_row + 1)
        sheet.add_chart(chart, chart_position)

        # Save the workbook
        book.save(path + name + "/Data.xlsx")
        return df


    # Helper function that generates an excel file from historic and predicted data
    # dct: dictionary giving giving an array of predicted values for every time key
    # x: pandas dataframe the ensemble was trained on (must have "time" + "Total Pax" columns)
    # time: int representing the number of months being extrapolated
    # key: list of strings that represent [type, airport, airline]
    # path: the string file path which the graphs are being saved to
    # seats: pandas dataframe of the future seat values (must have a "Total Seats" column)
    # name: string name of report that is being generated
    # weights: array of ints that decide how each model should be weighted
    # in the final average
    # INCLUDES OPTIMISTIC/PESSIMISTIC PREDICTIONS IF NOT USE save_numbers
    # USED FOR BACKTESTING IF YOU WANT TO EXTRAPOLATE USE save_numbers2 in Future

    def save_numbers2(self, dct, x, time, key, path, seats, name, weights):
        cls = self.__class__
        mean_dict = {key: np.average(value, weights=weights) for key, value in dct.items()}
        real_values = x[cls.target].values[time:]
        data = {
            "time": list(range(time, len(x["time"]))),
            "Type": key[0],
            "Airport": key[1],
            "Airline": key[2],
            "Actual Pax": real_values,
            "Mean Pred Pax": list(mean_dict.values()),
            "Percent Error": np.abs(100*((np.array(real_values) - np.array(list(mean_dict.values()))) / np.array(real_values))),
            "Total Seats": seats[time:],
            }
        df = pd.DataFrame(data)
        base_date = pd.Timestamp('2016-01-01') if cls.monthly else datetime(start_year, 1, 1)
        if cls.monthly:
            df['Date'] = df['time'].apply(lambda x: base_date + pd.DateOffset(months=x))
            df['Date'] = df['Date'].dt.strftime('%b %Y')
        else:
            df['Date'] = df['time'].apply(lambda x: base_date + pd.DateOffset(years=x))
            df['Date'] = df['Date'].dt.strftime('%Y')
        df.drop(columns = "time", inplace = True)
        df["Mean Pred Pax"] = df["Mean Pred Pax"].round().astype(int)
        df["Percent Error"] = df["Percent Error"].round(1).astype(str) + '%'
        df["Actual LF"] = get_lf(df["Actual Pax"], df["Total Seats"])
        df["Mean Pred LF"] = get_lf(df["Mean Pred Pax"], df["Total Seats"])
        # Rearrange the columns to place the date column first
        
        
        cols = ["Date", "Type", "Airport","Airline", "Actual Pax", "Mean Pred Pax", "Percent Error",
                "Total Seats", "Actual LF", "Mean Pred LF"] if self.monthly else \
                ["Date", "Type", "Airport","Airline", "Actual Pax", "Mean Pred Pax", "Percent Error"]
        df = df[cols]

        df.title = get_title(key)

        sheet_name = self.get_date(time) # One sheet per time interval

        # Load existing workbook or create a new one if it doesn't exist
        try:
            book = load_workbook(path + name + "/Data.xlsx")
            if sheet_name not in book.sheetnames:
                # Create the sheet if it doesn't exist
                book.create_sheet(sheet_name)
            with pd.ExcelWriter(path + name + "/Data.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                startrow = book[sheet_name].max_row  # Find the last row in the existing sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
        except FileNotFoundError:
            with pd.ExcelWriter(path + name + "/Data.xlsx", engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                startrow = 0  # Since this is a new sheet

        # Load workbook and select sheet to add graph
        book = load_workbook(path + name + "/Data.xlsx")
        sheet = book[sheet_name]

        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
            col_letter = col[0].column_letter  # Get the column letter
            sheet.column_dimensions[col_letter].width = 13

        # Create a line chart
        chart = LineChart()
        chart.title = get_title(key)[:-1]
        chart.style = 10
        chart.height = 6
        chart.width = 12
        chart.grouping = 'standard'
        chart.y_axis.title = "Predicted Passengers"
        chart.x_axis.title = cls.xlabel

        size = current_month - time
        # Iterate over all rows and set their height
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            sheet.row_dimensions[row[0].row].height = 150 / size if size < 10 else 15

        # Define data for the chart using only the newly added rows
        new_data_start_row = startrow + 1
        new_data_end_row = sheet.max_row
        data = Reference(sheet, min_col=5, min_row=new_data_start_row, max_col=6, max_row=new_data_end_row)
        categories = Reference(sheet, min_col=1, min_row=new_data_start_row + 1, max_row=new_data_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Set the minimum value for the y-axis
        min_value = min(df[["Mean Pred Pax", "Actual Pax"]].min())
        chart.y_axis.scaling.min = min_value

        # Add the chart to the sheet
        fst = "M" if self.monthly else "J"
        chart_position = fst + str(new_data_start_row + 1)
        sheet.add_chart(chart, chart_position)

        # Save the workbook
        book.save(path + name + "/Data.xlsx")
        return df
    

    # Returns a str that gives the date in MMM-YYYY format given a time inputted by the user
    # time = integer inputted by the user in PAARM.py which represents the number of months since current_month

    def get_date(self, time):
        year = ((time + 1) // 12) + 2016
        month = (time % 12) + 1
        date = datetime(int(year), int(month), 1)
        return date.strftime("%b - %Y")