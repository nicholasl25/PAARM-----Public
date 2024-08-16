"""# Old Helper Functions"""
# This file contains old code imported from Google Colab that is outdated and no longer used
# but is saved here in case it is ever needed again

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from statsmodels.tsa.statespace.sarimax import SARIMAX
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
from Classes.Helper import get_title, standard_dev
from openpyxl.chart import LineChart, Reference
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
from Classes.Helper import Pax_dict, check_count


""""#### Helper Function: lin_reg"""

# Returns a dataframe of the predicted values by a strict linear regression
# x: the independent variable (must be a series/array/column)
# y: the dependent variable (must be a series/array/column)
# time: the number of point the model is trained on


def lin_reg(x, y, time):
  x_train = x[:time]
  y_train = y[:time]
  x_test = x[time:]
  model = LinearRegression()
  model.fit(x_train, y_train)
  y_pred = model.predict(x_test)
  return pd.DataFrame(y_pred)

"""#### Helper Function: get_graph"""

# Exports and shows a formated graph of real vs. predicted values and MAPE
# USED FOR BACKTESTING NOT EXTRAPOLATING USE get_graph2 instead
# x: pandas column/series of independent variable (x)
# real: pandas column/series of real dependent variable
# time: the x value of where the prediction starts
# pred: pandas column/series of predicted dependent variable
# title: title of graph to be exported
# lst: key values that denote what airport/airline/type is being analyzed
# path: the directory that the graph will be saves to

def get_graph(x, real, pred, time, title, lst, path = "/content/drive/MyDrive/Port Authority2/time = 90/"):
    assert time <= len(x), "NICK SAYS: time is out of range of values"
    MAPE = np.mean(np.abs((real[time:].values - pred.values) / real[time:].values)) * 100
    plt.plot(x, real)
    plt.plot(x[time:], pred, label = "12 Month SARIMAX Model")
    plt.suptitle(title)
    plt.title("Mean Absolute Percentage Error (MAPE): " + str(round(MAPE, 2)) + "%", fontsize=10, color='red')
    plt.legend(['Actual Passenger Volumes', 'Predicted Passenger Volumes', "Training-Testing Split"], frameon=True, loc='lower left')
    plt.ylabel('Passenger Volume')
    plt.xlabel('Months Since Jan 2016')
    plt.axvline(x = time, color = 'black')
    plt.grid()
    directory1 = path + "Type = " + lst[0]
    directory2 = path + "Airport = " + lst[1]

    #outdated way of saving graphs to folder
"""
    directories = [directory1, directory2]
    for directory in directories:
      if not os.path.exists(directory):
          os.makedirs(directory)
    plt.savefig(directory1 +  "/" + lst[1] + "" + lst[3] + ".png")
    plt.savefig(directory2 +  "/" + lst[0] + "" + lst[3] + ".png") """

"""#### Helper Functions: get_graph2"""

# Exports and shows a formated graph of real vs. predicted values
# USED FOR EXTRAPOLATING NOT BACKTESTING USE get_graph instead
# x: pandas column/series of independent variable (x)
# real: pandas column/series of real dependent variable
# time: the x value of where the prediction starts
# pred: pandas column/series of predicted dependent variable
# title: title of graph to be exported
# path: the directory that the graph will be saves to

def get_graph2(x, real, pred, time, title, path):
    #MAPE = np.mean(np.abs((real[time:].values - pred.values) / real[time:].values)) * 100
    plt.plot(x, real)
    plt.plot((list(range(len(x), time))), pred, label = "12 Month SARIMAX Model")
    plt.suptitle(title)
    plt.title(str(time - len(x)) + " months into the future", color = 'red')
    plt.legend(['Actual Passenger Volumes', 'Predicted Passenger Volumes'], frameon=True, loc='lower left')
    plt.ylabel('Passenger Volume')
    plt.xlabel('Months Since Jan 2016')
    plt.axvline(x = len(x)-1, color = 'black')
    plt.grid()
    if path is not None:
      if not os.path.exists(path + title):
        os.makedirs(path + title)
      plt.savefig(path + title + "/Graph.png")

"""#### Helper Functions: get_pred"""

# Returns the prediction of a SARIMAX(order = (0,1,0), seasonal = (0,1,0,12)) model
# DOES NOT SUPPORT EXOGENOUS VARIABLES, IF NOT USE get_pred2 or get_pred3
# MUST HAVE time BEFORE END OF VALUES AND EXOG
# values: the ordered dataframe the prediction is being done on
# time: the number of rows of the dataframe the SARIMAX model is trained on
# target: the feature of the dataframe being predicted by the SARIMAX Model (default = "Total Pax")

def get_pred(values, time, target = "Total Pax"):
  assert time <= len(values[target]), "NICK SAYS: time is out of range of values"
  SARIMAX_model = SARIMAX(values[target][:time], order=(0, 1, 0), seasonal_order=(0, 1, 0, 12))
  model_fit = SARIMAX_model.fit(disp=False)
  # make prediction
  return model_fit.predict(time, len(values[target])-1)

"""#### Helper Function: get_pred2"""

# Returns an array of length two [first, second]
# First is the prediction of a SARIMAX(order = (0,1,0), seasonal = (0,1,0,12)) model
# Second is the trained model itself

# MUST HAVE time BEFORE END OF VALUES AND EXOG
# MUST HAVE EXOGENOUS VARIABLE(S), IF NOT USE get_pred
# values: the ordered dataframe the prediction is being done on
# time: the number of rows of the dataframe the SARIMAX model is trained on
# target: the feature of the dataframe being predicted by the SARIMAX Model (default = "Total Pax")
# exog: the exogenous variable used to make predictions


def get_pred2(values, time, exog, target = "Total Pax"):
  assert time <= len(values[target]), "NICK SAYS: time is out of range of values"
  assert len(exog) == len(values[target]), "NICK SAYS: The number of exogenous variables must equal the number of values"
  SARIMAX_model2 = SARIMAX(values[target][:time], exog = exog[:time],
                             order=(0, 1, 0), seasonal_order=(0, 1, 0, 12))
  SARIMAX_model_fit2 = SARIMAX_model2.fit(disp=False)
  # make prediction
  return [SARIMAX_model_fit2.predict(time, len(exog)-1, exog = exog[time:]), SARIMAX_model_fit2]

"""#### Helper Function: get_pred3"""

# Returns the prediction of a SARIMAX(order = (0,1,0), seasonal = (0,1,0,12)) model - Used to extrapolate into the future
# MUST HAVE EXOGENOUS VARIABLE(S), IF NOT USE get_pred
# values: the ordered dataframe the prediction is being done on
# exog1: the exogenous variables before the end of values
# exog2: the exogenous variables after the end of values
# target: the feature of the dataframe being predicted by the SARIMAX Model (default = "Total Pax")

def get_pred3(values, exog1, exog2, target = "Total Pax"):
  assert len(exog1) == len(values[target]), "NICK SAYS: The number of exogenous variables must equal the number of values"
  SARIMAX_model3 = SARIMAX(values[target], exog = exog1,
                             order=(0, 1, 0), seasonal_order=(0, 1, 0, 12))
  SARIMAX_model_fit3 = SARIMAX_model3.fit(disp=False)
  # make prediction
  return SARIMAX_model_fit3.predict(len(exog1), len(exog1)+len(exog2)-1, exog = exog2)

def save_numbers(dct, x, time, key, path, seats, weights):
  mean_dict = {key: np.average(value, weights=weights) for key, value in dct.items()}
  std_dict = {key: standard_dev(value, weights) for key, value in dct.items()}
  opt_dict = {key: np.mean(value) + np.std(value) for key, value in dct.items()}
  pess_dict = {key: np.mean(value) - np.std(value) for key, value in dct.items()}
  data = {
    "time": list(range(len(x), len(x) + time)),
    "Mean Pred Pax": list(mean_dict.values()),
    "Opt Pred Pax": list(opt_dict.values()),
    "Pes Pred Pax":  list(pess_dict.values()),
    "Total Seats": seats

    }
  df = pd.DataFrame(data)
  base_date = pd.Timestamp('2016-01-01')
  df['Date'] = df['time'].apply(lambda x: base_date + pd.DateOffset(months=x))
  df.drop(columns = "time", inplace = True)
  df['Date'] = df['Date'].dt.strftime('%b %Y')
  df["Mean Pred Pax"] = df["Mean Pred Pax"].round().astype(int)
  df["Opt Pred Pax"] = df["Opt Pred Pax"].round().astype(int)
  df["Pes Pred Pax"] = df["Pes Pred Pax"].round().astype(int)
  df['Mean Pred LF'] = (df['Mean Pred Pax'] / df['Total Seats']).round(3)
  df['Opt Pred LF'] = (df["Opt Pred Pax"] / df['Total Seats']).round(3)
  df['Pes Pred LF'] = (df["Pes Pred Pax"] / df['Total Seats']).round(3)
  # Rearrange the columns to place the date column first
  df = df[["Date", "Mean Pred Pax", "Opt Pred Pax", "Pes Pred Pax", "Total Seats", "Mean Pred LF", "Opt Pred LF", "Pes Pred LF"]]
  df.title = get_title(key)


  # Load existing workbook or create a new one if it doesn't exist
  try:
      book = load_workbook(path + "Sheet.xlsx")
      with pd.ExcelWriter(path + "Sheet.xlsx", engine='openpyxl', mode='a') as writer:
          df.to_excel(writer, sheet_name=get_title(key), index=False)
  except FileNotFoundError:
      with pd.ExcelWriter(path + "Sheet.xlsx", engine='openpyxl') as writer:
          df.to_excel(writer, sheet_name=get_title(key), index=False)

  # Load workbook and select sheet to add graph
  book = load_workbook(path + "Sheet.xlsx")
  sheet = book[get_title(key)]

  # Create a line chart
  chart = LineChart()
  chart.title = get_title(key)
  chart.style = 10
  chart.y_axis.title = "Predicted Passengers"
  chart.x_axis.title = "Months Since Jan 2016"

  # Define data for the chart
  data = Reference(sheet, min_col=2, min_row=1, max_col=4, max_row=len(df) + 1)
  categories = Reference(sheet, min_col=1, min_row=2, max_row=len(df) + 1)
  chart.add_data(data, titles_from_data=True)
  chart.set_categories(categories)

  # Set the minimum value for the y-axis
  min_value = min(df[["Mean Pred Pax", "Opt Pred Pax", "Pes Pred Pax"]].min())
  chart.y_axis.scaling.min = min_value

  # Add the chart to the sheet
  sheet.add_chart(chart, "K2")

  # Save the workbook
  book.save(path + "Sheet.xlsx")

  return df

# Outdated - not used
def get_numbers(pred, x, time, title, path):
  data = {
    "time": list(range(len(x), len(x) + time)),
    "Predicted Passengers": pred
  }
  df = pd.DataFrame(data)
  base_date = pd.Timestamp('2016-01-01')
  df['Date'] = df['time'].apply(lambda x: base_date + pd.DateOffset(months=x))
  df.drop(columns = "time", inplace = True)
  df['Date'] = df['Date'].dt.strftime('%B %Y')
  df['Predicted Passengers'] = df['Predicted Passengers'].round().astype(int)

  # Rearrange the columns to place the date column first
  df = df[['Date', 'Predicted Passengers']]
  df.title = title
  if path is not None:
    if not os.path.exists(path + title):
      os.makedirs(path + title)
    df.to_csv(path + title + "/Data.csv")
  return df


"""## ACT/PACF ANALYSIS: OUTDATED CODE USED IN EXPLOARTORY DATA ANALYSIS"""

# File for analyzing the autocorrelation function (ACF) and partial autocorrelation function (PACF)
# of Pax data and generating a graph. (These functions are useful as they help us determine
# what degree autoregression/moving average are statistically significant)


path = "/Users/nlombardo/Desktop/PAARM/Reports/ACF/"
target = "Total Pax"

os.makedirs(path, exist_ok=True)

for key in Pax_dict:
    if check_count(Pax_dict[key]):
        data = np.diff(Pax_dict[key][target])
        plt.figure(figsize=(12, 6))

        # Plot ACF
        plt.subplot(1, 2, 1)
        plot_acf(data, lags=20, ax=plt.gca())
        plt.title('Autocorrelation Function (ACF)')

        # Plot PACF
        plt.subplot(1, 2, 2)
        plot_pacf(data, lags=20, ax=plt.gca(), method='yw')
        plt.title('Partial Autocorrelation Function (PACF)')

        plt.suptitle(key)
        plt.tight_layout()
        plt.savefig(path + key[0] + " " + key[1] + " "  + key[2])
        plt.close()
